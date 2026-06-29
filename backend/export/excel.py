"""Excel export for matching reports using openpyxl (DASH-02, D-06, D-09).

Generates a single-sheet .xlsx workbook with candidate data columns.
openpyxl handles Unicode natively, no special font setup needed for
Chinese characters (Assumption A2 from research).
"""

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_match_report_excel(
    candidates: list[dict[str, Any]],
    output_path: str,
    sheet_title: str = "匹配结果",
) -> str:
    """Generate a single-sheet Excel matching report (D-09).

    Columns (per D-09):
    A: 候选人姓名, B: 匹配度, C: 匹配技能, D: 缺失技能,
    E: 审核状态, F: 备注

    Args:
        candidates: List of MatchResult dicts with keys:
            candidate_name, match_score, matched_skills, missing_skills,
            should_proceed, recommendation.
        output_path: Absolute path where the .xlsx will be written.
        sheet_title: Title for the worksheet tab.

    Returns:
        The output_path for chaining.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # ── Headers (D-09) ──
    headers = ["候选人姓名", "匹配度", "匹配技能", "缺失技能", "审核状态", "备注"]
    ws.append(headers)

    # ── Style headers ──
    header_font = Font(name="Noto Sans CJK SC", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    for c in candidates:
        status = "通过" if c.get("should_proceed") else (
            "驳回" if c.get("should_proceed") is False else "待审"
        )
        ws.append([
            c.get("candidate_name", ""),
            c.get("match_score", 0),
            ", ".join(c.get("matched_skills", []) or []),
            ", ".join(c.get("missing_skills", []) or []),
            status,
            c.get("recommendation", ""),
        ])

    # Style data rows with border
    data_font = Font(name="Noto Sans CJK SC", size=10)
    data_align = Alignment(vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # ── Column widths ──
    ws.column_dimensions["A"].width = 18   # name
    ws.column_dimensions["B"].width = 10   # score
    ws.column_dimensions["C"].width = 40   # matched skills
    ws.column_dimensions["D"].width = 40   # missing skills
    ws.column_dimensions["E"].width = 12   # status
    ws.column_dimensions["F"].width = 30   # notes

    wb.save(output_path)
    return output_path
